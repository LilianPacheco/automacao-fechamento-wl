from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def text_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


def piece_key(value: object) -> str:
    compact = text_key(value)
    match = re.fullmatch(r"([A-Z]+)0*(\d+)([A-Z]?)", compact)
    if match:
        return f"{match.group(1)}{int(match.group(2))}{match.group(3)}"
    return compact


def number(value: object) -> float | None:
    if value in (None, "", "DIG.VOLUME"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        raw = str(value).strip()
        return float(raw.replace(".", "").replace(",", ".") if "," in raw else raw)
    except ValueError:
        return None


def close_number(left: object, right: object, tolerance: float = 0.002) -> bool:
    a, b = number(left), number(right)
    return a is not None and b is not None and abs(a - b) <= tolerance


def workbook_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook[sheet_name]
    rows: list[dict[str, object]] = []
    for row_number in range(7, sheet.max_row + 1):
        date = sheet.cell(row_number, 2).value
        if not isinstance(date, datetime):
            continue
        prefix = str(sheet.cell(row_number, 5).value or "").strip()
        suffix = str(sheet.cell(row_number, 6).value or "").strip()
        piece = f"{prefix}-{suffix}" if prefix and suffix else prefix
        dimensions = str(sheet.cell(row_number, 7).value or "").strip()
        section_match = re.search(r"\b\d{1,3}\s*[X×]\s*\d{1,3}(?:/\d{1,3})?\b", dimensions, re.I)
        length_matches = re.findall(r"\d+[,.]\d{3}", dimensions)
        rows.append({
            "row": row_number,
            "date": date.strftime("%d/%m/%Y"),
            "type": sheet.cell(row_number, 1).value or "",
            "work": sheet.cell(row_number, 3).value or "",
            "quantity": sheet.cell(row_number, 4).value,
            "piece": piece,
            "section": section_match.group(0).replace("×", "X").replace(" ", "") if section_match else "",
            "length": length_matches[-1] if length_matches else "",
            "volume": sheet.cell(row_number, 8).value,
        })
    workbook.close()
    return rows


def review_rows(path: Path) -> list[dict[str, object]]:
    loaded = json.loads(path.read_text(encoding="utf-8"))
    groups = loaded.values() if isinstance(loaded, dict) else [loaded]
    return [row for group in groups for row in (group if isinstance(group, list) else [group]) if isinstance(row, dict)]


def candidate_score(review: dict[str, object], manual: dict[str, object]) -> int:
    score = 0
    if piece_key(review.get("piece")) and piece_key(review.get("piece")) == piece_key(manual.get("piece")):
        score += 8
    if text_key(review.get("section")) and text_key(review.get("section")) == text_key(manual.get("section")):
        score += 4
    if close_number(review.get("length"), manual.get("length")):
        score += 3
    if close_number(review.get("unit_volume"), manual.get("volume")):
        score += 3
    if text_key(review.get("work")) and text_key(review.get("work")) in text_key(manual.get("work")):
        score += 2
    if text_key(review.get("type_name")) == text_key(manual.get("type")):
        score += 1
    return score


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--review", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--sheet", default="2ª quinz.agosto")
    args = parser.parse_args()

    manual = workbook_rows(args.workbook, args.sheet)
    review = review_rows(args.review)
    automatic = [row for row in review if not row.get("warnings")]
    matched: list[tuple[dict[str, object], dict[str, object], int]] = []
    unmatched = []
    for row in automatic:
        candidates = [item for item in manual if item["date"] == row.get("message_date")]
        scored = sorted(((candidate_score(row, item), item) for item in candidates), key=lambda pair: pair[0], reverse=True)
        if not scored or scored[0][0] < 8:
            unmatched.append(row)
            continue
        matched.append((row, scored[0][1], scored[0][0]))

    checks = Counter()
    correct = Counter()
    for row, expected, _ in matched:
        comparisons = {
            "tipo": text_key(row.get("type_name")) == text_key(expected.get("type")),
            "obra": text_key(row.get("work")) in text_key(expected.get("work")) or text_key(expected.get("work")) in text_key(row.get("work")),
            "peça": piece_key(row.get("piece")) == piece_key(expected.get("piece")),
            "seção": text_key(row.get("section")) == text_key(expected.get("section")),
            "comprimento": close_number(row.get("length"), expected.get("length")),
            "volume": close_number(row.get("unit_volume"), expected.get("volume")),
        }
        for field, is_correct in comparisons.items():
            expected_field = {
                "tipo": expected.get("type"),
                "obra": expected.get("work"),
                "peça": expected.get("piece"),
                "seção": expected.get("section"),
                "comprimento": expected.get("length"),
                "volume": number(expected.get("volume")),
            }[field]
            if expected_field in (None, ""):
                continue
            checks[field] += 1
            correct[field] += int(is_correct)

    result = {
        "linhas_manuais_datadas": len(manual),
        "fotos_ou_entradas_lidas": len(review),
        "entradas_sem_pendencia": len(automatic),
        "entradas_com_pendencia": len(review) - len(automatic),
        "automaticas_identificadas_no_manual": len(matched),
        "automaticas_sem_correspondencia_segura": len(unmatched),
        "precisao_por_campo_nas_correspondencias": {
            field: {
                "corretos": correct[field],
                "avaliados": checks[field],
                "percentual": round(correct[field] / checks[field] * 100, 1) if checks[field] else None,
            }
            for field in checks
        },
        "observacao": "Correspondência exige mesma data e mesmo código de peça; estacas sem código individual não entram nesta precisão.",
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
