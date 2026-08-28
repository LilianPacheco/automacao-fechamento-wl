from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\lilia\Downloads\FECHAMENTOS PRELOG\MEDIÇÕES AWL - 2026.xlsx")


def prepare(path: Path = SOURCE, output: Path | None = None) -> tuple[Path, list[str]]:
    if not path.exists():
        raise FileNotFoundError(path)
    backup_dir = (path.parent / "Backups Fechamento") if output is None else Path(__file__).resolve().parents[2] / "Backups Fechamento"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{path.stem}_antes_agosto_{datetime.now():%Y-%m-%d_%H-%M-%S}{path.suffix}"
    shutil.copy2(path, backup)

    workbook = load_workbook(path, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    template = workbook["2ª quinz.julho"]
    created: list[str] = []
    for name, start, end in (("1ª quinz.agosto", 1, 15), ("2ª quinz.agosto", 16, 31)):
        if name in workbook.sheetnames:
            del workbook[name]
        sheet = workbook.copy_worksheet(template)
        sheet.title = name
        sheet["C3"] = datetime(2026, 8, start)
        sheet["E3"] = datetime(2026, 8, end)
        sheet["C3"].number_format = "dd/mm/yyyy"
        sheet["E3"].number_format = "dd/mm/yyyy"
        # Keep formulas, formatting and the DIG.VOLUME input marker, but clear
        # all source fields so the new fortnight starts empty.
        for row in range(7, 304):
            for col in (1, 2, 3, 4, 5, 6, 7, 10):
                sheet.cell(row, col).value = None
            sheet.cell(row, 8).value = "DIG.VOLUME"
        created.append(name)
    destination = output or path
    workbook.save(destination)
    workbook.close()
    return backup, created


if __name__ == "__main__":
    output = Path(__import__("sys").argv[1]) if len(__import__("sys").argv) > 1 else None
    backup_path, sheets = prepare(output=output)
    print(f"output={output or SOURCE}")
    print(f"backup={backup_path}")
    print("sheets=" + ",".join(sheets))
