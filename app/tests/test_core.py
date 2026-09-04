from __future__ import annotations

import tempfile
import unittest
import json
import base64
from datetime import date, datetime
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw

from wl_fechamento.config import load_configuration, save_configuration
from wl_fechamento.label_parser import parse_label_text
from wl_fechamento.ocr_service import _orange_label_crop, score_ocr_text
from wl_fechamento.chrome_bridge import BRIDGE_PORT, ChromeBridge, load_saved_whatsapp_session
from wl_fechamento.models import AppConfiguration, PeriodSelection
from wl_fechamento.stake_parser import parse_stake_text
from wl_fechamento.whatsapp_service import (
    WhatsAppProbeResult,
    _parse_probe_output,
    merge_period_results,
    restrict_result_to_period,
)
from wl_fechamento.label_parser import normalize_type, parse_document_text, parse_label_text
from wl_fechamento.grouping_service import ConsolidatedRow, group_approved_drafts
from wl_fechamento.review_service import (
    _apply_message_consensus,
    _apply_message_quantity,
    _sanitize_duplicate_measurements,
)
from wl_fechamento.workbook_writer_service import rows_to_payload, write_approved_rows
from wl_fechamento.workbook_service import create_backup, validate_workbook


class PeriodTests(unittest.TestCase):
    def test_second_fortnight_uses_last_day_of_month(self) -> None:
        period = PeriodSelection(2026, 2, 2)
        self.assertEqual(period.start_date.isoformat(), "2026-02-16")
        self.assertEqual(period.end_date.isoformat(), "2026-02-28")
        self.assertEqual(period.sheet_name, "2ª quinz.fevereiro")

    def test_first_fortnight(self) -> None:
        period = PeriodSelection(2026, 7, 1)
        self.assertEqual(period.start_date.isoformat(), "2026-07-01")
        self.assertEqual(period.end_date.isoformat(), "2026-07-15")

    def test_review_key_is_unique_per_fortnight(self) -> None:
        self.assertEqual(
            PeriodSelection(2026, 8, 2).review_key,
            "2026-08_2a-quinzena",
        )

    def test_probe_result_discards_evidence_from_other_fortnight(self) -> None:
        result = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "period_scan_complete": True,
            "start_date": "02/08/2026",
            "evidences": [
                {"message_id": "old", "message_date": "02/08/2026", "image_count": 1},
                {"message_id": "current", "message_date": "18/08/2026", "image_count": 1, "quantity_hint": 21},
            ],
            "captured_attachments": [
                {"message_id": "old", "filename": "old.jpg", "path": "old.jpg", "size": 1, "sha256": "a"},
                {"message_id": "current", "filename": "current.jpg", "path": "current.jpg", "size": 1, "sha256": "b"},
            ],
        })

        filtered = restrict_result_to_period(
            result,
            date(2026, 8, 16),
            date(2026, 8, 31),
        )

        self.assertEqual([item.message_id for item in filtered.evidences], ["current"])
        self.assertEqual(
            [item.message_id for item in filtered.captured_attachments],
            ["current"],
        )
        self.assertEqual(filtered.start_date, "16/08/2026")
        self.assertTrue(filtered.start_date_found)
        self.assertEqual(filtered.evidences[0].quantity_hint, 21)
        self.assertIn("primeira evidência reconhecida em 18/08/2026", filtered.message)

    def test_partial_period_evidence_never_releases_review(self) -> None:
        result = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "period_scan_complete": False,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "late-album",
                "message_date": "27/08/2026",
                "image_count": 5,
            }],
            "incomplete_albums": [{
                "message_id": "late-album",
                "expected": 5,
                "captured": 0,
            }],
        })

        filtered = restrict_result_to_period(
            result,
            date(2026, 8, 16),
            date(2026, 8, 31),
        )

        self.assertFalse(filtered.start_date_found)
        self.assertFalse(filtered.period_scan_complete)
        self.assertEqual(len(filtered.incomplete_albums), 1)

    def test_complete_current_read_uses_older_attempts_only_for_attachments(self) -> None:
        primary = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "period_scan_complete": True,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "late",
                "message_date": "23/08/2026",
                "image_count": 1,
            }],
        })
        earlier = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "early",
                "message_date": "17/08/2026",
                "image_count": 1,
            }],
        })

        merged = merge_period_results(
            primary, [earlier], date(2026, 8, 16), date(2026, 8, 31)
        )
        filtered = restrict_result_to_period(
            merged, date(2026, 8, 16), date(2026, 8, 31)
        )

        self.assertEqual(
            [item.message_date for item in filtered.evidences],
            ["23/08/2026"],
        )
        self.assertTrue(filtered.period_scan_complete)
        self.assertTrue(filtered.start_date_found)
        self.assertIn(
            "primeira evidência reconhecida em 23/08/2026",
            filtered.message,
        )
        self.assertEqual(len(filtered.incomplete_albums), 1)

    def test_partial_current_read_can_merge_older_inventory(self) -> None:
        primary = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "period_scan_complete": False,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "late",
                "message_date": "23/08/2026",
                "image_count": 0,
            }],
        })
        earlier = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "early",
                "message_date": "17/08/2026",
                "image_count": 0,
            }],
        })

        merged = merge_period_results(
            primary, [earlier], date(2026, 8, 16), date(2026, 8, 31)
        )

        self.assertEqual(
            [item.message_date for item in merged.evidences],
            ["17/08/2026", "23/08/2026"],
        )
        self.assertFalse(merged.period_scan_complete)

    def test_only_primary_read_can_prove_complete_period_scan(self) -> None:
        partial = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": False,
            "period_scan_complete": False,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "late",
                "message_date": "23/08/2026",
                "image_count": 0,
            }],
        })
        old_complete = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "period_scan_complete": True,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "early",
                "message_date": "17/08/2026",
                "image_count": 0,
            }],
        })

        merged = merge_period_results(
            partial, [old_complete], date(2026, 8, 16), date(2026, 8, 31)
        )

        self.assertFalse(merged.period_scan_complete)
        self.assertFalse(merged.start_date_found)


    def test_photo_caption_quantity_replaces_default_piece_count(self) -> None:
        result = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "start_date": "16/08/2026",
            "evidences": [{
                "message_id": "photo-21",
                "message_date": "20/08/2026",
                "image_count": 1,
                "quantity_hint": 21,
            }],
        })
        draft = parse_label_text(
            "Obra: TESTE Produto: PILAR Secao: 40X40 "
            "Comprimento: 8,000 Peca: PH-1 Vol. (m3): 1,250",
            message_id="photo-21",
            message_date="20/08/2026",
        )

        _apply_message_quantity([draft], result, "photo-21")

        self.assertEqual(draft.quantity, 21)


class StakeParserTests(unittest.TestCase):
    def test_approved_example(self) -> None:
        entry = parse_stake_text("16x10=600")
        self.assertEqual(entry.type_name, "ESTACA")
        self.assertEqual(entry.piece, "16")
        self.assertEqual(entry.quantity, 6000)


class OcrCropTests(unittest.TestCase):
    def test_dim_orange_frame_is_detected(self) -> None:
        image = Image.new("RGB", (900, 1600), (105, 105, 105))
        draw = ImageDraw.Draw(image)
        draw.rectangle((160, 420, 700, 740), outline=(149, 71, 22), width=14)

        crop = _orange_label_crop(image)

        self.assertIsNotNone(crop)
        self.assertGreater(crop.width, 450)
        self.assertGreater(crop.height, 250)

    def test_red_handwriting_is_not_joined_to_orange_label(self) -> None:
        image = Image.new("RGB", (800, 900), "#d8d8d8")
        draw = ImageDraw.Draw(image)
        draw.rectangle((120, 90, 560, 340), outline=(238, 115, 24), width=18)
        draw.line((160, 650, 650, 500), fill=(185, 36, 30), width=24)

        crop = _orange_label_crop(image)

        self.assertIsNotNone(crop)
        self.assertLess(crop.height, 450)
        self.assertGreater(crop.width, 450)
        self.assertGreater(crop.height, 250)

    def test_extended_example_uses_last_two_numbers(self) -> None:
        entry = parse_stake_text("16x16x8+100")
        self.assertEqual(entry.piece, "16")
        self.assertEqual(entry.multiplier, 8)
        self.assertEqual(entry.base_value, 100)
        self.assertEqual(entry.quantity, 800)
        self.assertEqual(entry.dimensions, "")
        self.assertIsNone(entry.unit_volume)

    def test_spaces_and_multiplication_symbol(self) -> None:
        entry = parse_stake_text(" 16 × 10 = 600 ")
        self.assertEqual(entry.quantity, 6000)

    def test_incomplete_text_is_rejected(self) -> None:
        with self.assertRaises(ValueError):
            parse_stake_text("16x10")


class LabelParserTests(unittest.TestCase):
    def test_visible_ph14_label_recovers_reordered_ocr_fields(self):
        raw = (
            "Secac: 40X60 24,520 Compr{mento (m): Peso 14822,26 "
            "pH—14 Peca: Prellog VENTISOL - GALPAO VERTICAL 282 "
            "Obra: 2301 Sigla: Produto: PILAR - RETANGULAR "
            "45634 5,834 Voj. (m3):"
        )
        result = parse_label_text(raw, message_date="02/08/2026")
        self.assertEqual(result.piece, "PH-14")
        self.assertEqual(result.section, "40X60")
        self.assertEqual(result.length, "24,520")
        self.assertEqual(result.unit_volume, 5.834)
        self.assertEqual(result.product, "PILAR")
        self.assertEqual(result.type_name, "PILAR")
        self.assertEqual(result.warnings, [])

    def test_romaneio_creates_one_draft_per_piece(self) -> None:
        drafts = parse_document_text(
            """
            ROMANEIO DE CARGAS | 0104 AMPLIACAO ALUMINIO SJ Obra
            PRELLOG PREFABRICA PRELLOG PREFABRICA PL120 PL124 Total:
            E=10 E=10 11,290 9,520 20,810 m Volume 1,350 1,140 2,490 m3
            Peso 3177,01 2678,80 PAINEL PAINEL MACICO MACICO
            """,
            message_id="romaneio-1",
            message_date="26/07/2026",
        )
        self.assertEqual([item.piece for item in drafts], ["PL120", "PL124"])
        self.assertEqual([item.section for item in drafts], ["E=10", "E=10"])
        self.assertEqual([item.length for item in drafts], ["11,290", "9,520"])
        self.assertEqual([item.unit_volume for item in drafts], [1.35, 1.14])
        self.assertEqual([item.type_name for item in drafts], ["PAINEL", "PAINEL"])

    def test_stake_delivery_uses_total_meters_as_quantity(self) -> None:
        drafts = parse_document_text(
            """
            ROMANEIO DE ENTREGA N 355/2026
            DESTINATARIO OBRA: AUTOLABOR INOVACAO E TECNOLOGIA PALHOCA
            Caminhao GGU1G61 Carreta RLGOA21 Motorista ANDRE
            Peca ESTACA ESTACA LUVA Dimensao 20x20 20x20 20x20
            Quantidade 4 2 4 Comprimento 8,00 10,00
            Metros 32 20 52 Peso 0,80 1,00 5,20
            """,
            message_id="entrega-355",
            message_date="06/08/2026",
        )

        self.assertEqual(len(drafts), 1)
        draft = drafts[0]
        self.assertEqual(draft.work, "AUTOLABOR INOVACAO E TECNOLOGIA PALHOCA")
        self.assertEqual(draft.product, "ESTACA")
        self.assertEqual(draft.type_name, "ESTACA")
        self.assertEqual(draft.piece, "20X20")
        self.assertEqual(draft.section, "20X20")
        self.assertEqual(draft.dimensions, "20X20")
        self.assertEqual(draft.quantity, 52)
        self.assertEqual(draft.length, "")
        self.assertIsNone(draft.unit_volume)
        self.assertEqual(draft.warnings, [])

    SAMPLE = (
        "Obra: AMPLIACAO ALUMINIO SJ Sigla: 0104 "
        "Produto: PAINEL - MACICO Secao: E=10 "
        "Comprimento (m): 11,110 Peso (kg): 3126,45 "
        "Peca: PL1 Vol. (m3): 1,330"
    )

    def test_reads_confirmed_label_fields(self) -> None:
        result = parse_label_text(self.SAMPLE)
        self.assertEqual(result.work, "AMPLIACAO ALUMINIO SJ")
        self.assertEqual(result.acronym, "0104")
        self.assertEqual(result.product, "PAINEL - MACICO")
        self.assertEqual(result.section, "E=10")
        self.assertEqual(result.length, "11,110")
        self.assertEqual(result.unit_volume, "1,330")
        self.assertEqual(result.piece_code, "PL1")
        self.assertEqual(result.issues, [])

    def test_ambiguous_piece_is_not_guessed(self) -> None:
        result = parse_label_text(self.SAMPLE.replace("Peca: PL1", "PLI Peca: 1219040"))
        self.assertEqual(result.piece_code, "")
        self.assertEqual(result.piece_candidate, "PLI")
        self.assertTrue(any("Confirmar código" in issue for issue in result.issues))

    def test_ocr_score_prefers_field_rich_text(self) -> None:
        self.assertGreater(score_ocr_text(self.SAMPLE), score_ocr_text("POCO X5 18/07/2026"))


class ConfigurationTests(unittest.TestCase):
    def test_configuration_roundtrip(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "config.json"
            original = AppConfiguration("C:/teste.xlsx", 2026, 7, 2)
            save_configuration(original, path)
            loaded = load_configuration(path)
            self.assertEqual(loaded, original)


class WhatsAppProbeTests(unittest.TestCase):
    def test_saved_session_is_recovered_from_local_images(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            message_id = "album-AAA-BBB-2"
            snapshot = {
                "connected": True,
                "group_found": True,
                "start_date_found": False,
                "start_date": "16/07/2026",
                "visible_images": 2,
                "evidences": [{
                    "message_id": message_id,
                    "message_date": "26/07/2026",
                    "image_count": 2,
                }],
            }
            (directory / "sessao_whatsapp.json").write_text(
                json.dumps(snapshot), encoding="utf-8"
            )
            (directory / f"{message_id}_hash_foto_1.jpg").write_bytes(b"one")
            (directory / f"{message_id}_hash_foto_2.jpg").write_bytes(b"two")
            result = load_saved_whatsapp_session(directory)
            self.assertEqual(len(result.captured_attachments), 2)
            self.assertEqual(result.captured_attachments[0].message_id, message_id)
            self.assertEqual(result.incomplete_albums, [])

    def test_structured_evidence_is_loaded(self) -> None:
        result = WhatsAppProbeResult.from_dict({
            "connected": True,
            "group_found": True,
            "start_date_found": True,
            "start_date": "16/07/2026",
            "evidences": [{
                "message_id": "message-1",
                "message_date": "17/07/2026",
                "message_time": "16:35",
                "sender": "Wilian",
                "image_count": 0,
                "pdf_names": ["Carga 115.pdf"],
                "stake_text": "",
                "has_ok": True,
            }],
            "incomplete_albums": [{
                "message_id": "album-1",
                "expected": 10,
                "captured": 4,
            }],
        })
        self.assertEqual(len(result.evidences), 1)
        self.assertEqual(result.evidences[0].pdf_names, ["Carga 115.pdf"])
        self.assertEqual(result.evidences[0].kind_label, "1 PDF")
        self.assertTrue(result.evidences[0].has_ok)
        self.assertEqual(result.incomplete_albums[0].expected, 10)
        self.assertEqual(result.incomplete_albums[0].captured, 4)

    def test_reads_last_result_line(self) -> None:
        output = "\n".join([
            '{"kind":"progress","stage":"connected"}',
            (
                '{"kind":"result","connected":true,"group_found":true,'
                '"start_date_found":true,"start_date":"16/07/2026",'
                '"load_attempts":2,"sync_waits":1,"sync_in_progress":false,'
                '"visible_images":9,"visible_pdfs":1,'
                '"ok_reactions":2,"stake_messages":["16×10=600"],'
                '"message":"Histórico comprovado."}'
            ),
        ])
        result = _parse_probe_output(output)
        self.assertTrue(result.connected)
        self.assertTrue(result.group_found)
        self.assertTrue(result.start_date_found)
        self.assertEqual(result.start_date, "16/07/2026")
        self.assertEqual(result.load_attempts, 2)
        self.assertEqual(result.sync_waits, 1)
        self.assertFalse(result.sync_in_progress)
        self.assertEqual(result.stake_messages, ["16×10=600"])

    def test_rejects_output_without_result(self) -> None:
        with self.assertRaises(RuntimeError):
            _parse_probe_output('{"kind":"progress","stage":"connected"}')

    def test_chrome_bridge_accepts_only_current_session(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            bridge = ChromeBridge(
                PeriodSelection(2026, 7, 2).start_date,
                PeriodSelection(2026, 7, 2).end_date,
                attachment_root=Path(temporary),
            )
            bridge.start()
            try:
                with urlopen(f"http://127.0.0.1:{BRIDGE_PORT}/config") as response:
                    config = json.loads(response.read().decode("utf-8"))
                self.assertEqual(config["start_label"], "16/07/2026")
                self.assertTrue(config["extension_version"])

                payload = {
                    "session_id": config["session_id"],
                    "token": config["token"],
                    "final": True,
                    "connected": True,
                    "group_found": True,
                    "start_date_found": True,
                    "start_date": config["start_label"],
                    "message": "Histórico comprovado.",
                }
                request = Request(
                    f"http://127.0.0.1:{BRIDGE_PORT}/result",
                    data=json.dumps(payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urlopen(request) as response:
                    self.assertEqual(response.status, 200)
                result = bridge.wait_for_result(timeout_seconds=1)
                self.assertTrue(result.start_date_found)
            finally:
                bridge.close()

    def test_chrome_bridge_ignores_final_result_from_wrong_profile(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            bridge = ChromeBridge(
                PeriodSelection(2026, 7, 2).start_date,
                PeriodSelection(2026, 7, 2).end_date,
                attachment_root=Path(temporary),
            )
            bridge.start()
            try:
                with urlopen(f"http://127.0.0.1:{BRIDGE_PORT}/config") as response:
                    config = json.loads(response.read().decode("utf-8"))

                def post_result(group_found: bool, message: str) -> None:
                    payload = {
                        "session_id": config["session_id"],
                        "token": config["token"],
                        "final": True,
                        "connected": True,
                        "group_found": group_found,
                        "start_date_found": group_found,
                        "start_date": config["start_label"],
                        "message": message,
                    }
                    request = Request(
                        f"http://127.0.0.1:{BRIDGE_PORT}/result",
                        data=json.dumps(payload).encode("utf-8"),
                        headers={"Content-Type": "application/json"},
                        method="POST",
                    )
                    with urlopen(request) as response:
                        self.assertEqual(response.status, 200)

                post_result(False, "Perfil pessoal sem o grupo.")
                self.assertFalse(bridge.result_ready.is_set())

                post_result(True, "Perfil AWL localizado.")
                result = bridge.wait_for_result(timeout_seconds=1)
                self.assertTrue(result.group_found)
                self.assertEqual(result.message, "Perfil AWL localizado.")
            finally:
                bridge.close()

    def test_chrome_bridge_stores_authorized_attachment(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            bridge = ChromeBridge(
                PeriodSelection(2026, 7, 2).start_date,
                PeriodSelection(2026, 7, 2).end_date,
                attachment_root=Path(temporary),
            )
            bridge.start()
            try:
                with urlopen(f"http://127.0.0.1:{BRIDGE_PORT}/config") as response:
                    config = json.loads(response.read().decode("utf-8"))
                content = b"small-image-content"
                payload = {
                    "session_id": config["session_id"],
                    "token": config["token"],
                    "message_id": "message-1",
                    "filename": "foto.jpeg",
                    "mime_type": "image/jpeg",
                    "base64": base64.b64encode(content).decode("ascii"),
                }
                request = Request(
                    f"http://127.0.0.1:{BRIDGE_PORT}/attachment",
                    data=json.dumps(payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urlopen(request) as response:
                    self.assertEqual(response.status, 200)
                self.assertEqual(len(bridge.attachments), 1)
                saved = Path(bridge.attachments[0]["path"])
                self.assertEqual(saved.read_bytes(), content)
                final_payload = {
                    "session_id": config["session_id"],
                    "token": config["token"],
                    "final": True,
                    "connected": True,
                    "group_found": True,
                    "start_date_found": True,
                    "start_date": "16/07/2026",
                }
                final_request = Request(
                    f"http://127.0.0.1:{BRIDGE_PORT}/result",
                    data=json.dumps(final_payload).encode("utf-8"),
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urlopen(final_request) as response:
                    self.assertEqual(response.status, 200)
                parsed = bridge.wait_for_result(timeout_seconds=1)
                self.assertEqual(parsed.captured_attachments[0].sha256, bridge.attachments[0]["sha256"])
                snapshot_path = Path(temporary) / "sessao_whatsapp.json"
                self.assertTrue(snapshot_path.exists())
                snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
                self.assertTrue(snapshot["final"])
                self.assertEqual(len(snapshot["captured_attachments"]), 1)
                self.assertEqual(
                    snapshot["captured_attachments"][0]["sha256"],
                    bridge.attachments[0]["sha256"],
                )
            finally:
                bridge.close()


class LabelParserTests(unittest.TestCase):
    def test_visible_ph14_label_recovers_reordered_ocr_fields(self):
        raw = (
            "Secac: 40X60 24,520 Compr{mento (m): Peso 14822,26 "
            "pH—14 Peca: Prellog VENTISOL - GALPAO VERTICAL 282 "
            "Obra: 2301 Sigla: Produto: PILAR - RETANGULAR "
            "45634 5,834 Voj. (m3):"
        )
        result = parse_label_text(raw, message_date="02/08/2026")
        self.assertEqual(result.piece, "PH-14")
        self.assertEqual(result.section, "40X60")
        self.assertEqual(result.length, "24,520")
        self.assertEqual(result.unit_volume, 5.834)
        self.assertEqual(result.product, "PILAR")
        self.assertEqual(result.type_name, "PILAR")
        self.assertEqual(result.warnings, [])

    def test_prellog_label_is_parsed_for_temporary_review(self) -> None:
        draft = parse_label_text(
            """
            Obra: AMPLIACAO ALUMINIO SJ
            Sigla: 0104
            Produto: PAINEL - MACICO
            Secao: E=10
            Comprimento (m): 11,110
            Peso (kg): 3126,45
            Peca: PL1
            Vol. (m3): 1,330
            """,
            message_date="25/07/2026",
        )
        self.assertEqual(draft.type_name, "PAINEL")
        self.assertEqual(draft.piece, "PL1")
        self.assertEqual(draft.dimensions, "E=10 11,110")
        self.assertEqual(draft.unit_volume, 1.33)
        self.assertEqual(draft.status, "PRONTO PARA REVISÃO")

    def test_piece_before_label_and_qr_number_do_not_corrupt_fields(self) -> None:
        draft = parse_label_text(
            "Obra: MKM GALPAO COMERCIAL Sigla: 0226 Produto: PILAR - RETANGULAR "
            "Secao: 30X50 Comprimento (m): 15,750 Peso (kg): 5116,17 "
            "pp-11 peca : Vol. 1619573 (mm: 2,030",
            message_date="26/07/2026",
        )
        self.assertEqual(draft.piece, "PP-11")
        self.assertEqual(draft.unit_volume, 2.03)
        self.assertEqual(draft.type_name, "PILAR")

    def test_explicit_tokens_survive_missing_ocr_labels(self) -> None:
        draft = parse_label_text(
            "VENTISOL - GALPAO VERTICAL 282 2301 VIGA - RETANGULAR "
            "30X80 Comprimento (m): 7,390 VPR-2 Vol. (m3): 1,551",
            message_date="19/07/2026",
        )
        self.assertEqual(draft.work, "VENTISOL - GALPAO VERTICAL 282")
        self.assertEqual(draft.product, "VIGA 6,1 ATÉ 8,9m")
        self.assertEqual(draft.section, "30X80")
        self.assertEqual(draft.piece, "VPR-2")

    def test_small_label_values_survive_when_ocr_drops_field_names(self) -> None:
        draft = parse_label_text(
            """
            Pmitog
            VENTINOL - ALPAO VERTICAL 2
            2301
            PILAR
            NETANULALAN
            4CX60
            13273,63
            21,800
            5.284
            PH-15
            """,
            message_date="02/08/2026",
        )
        self.assertEqual(draft.work, "VENTISOL - GALPAO VERTICAL 282")
        self.assertEqual(draft.product, "PILAR")
        self.assertEqual(draft.section, "40X60")
        self.assertEqual(draft.length, "21,800")
        self.assertEqual(draft.unit_volume, 5.284)
        self.assertEqual(draft.piece, "PH-15")
        self.assertEqual(draft.warnings, [])

    def test_piece_family_recovers_product_without_confusing_ph_and_pp(self) -> None:
        ph = parse_label_text(
            "VENTISOL - GALPAO VERTICAL 282 PELAR 40X60 21,800 5,284 PH-16",
            message_date="02/08/2026",
        )
        pp = parse_label_text(
            "VENTISOL - GALPAO VERTICAL 282 50X50 18,500 3,100 PP-16",
            message_date="02/08/2026",
        )
        self.assertEqual((ph.piece, ph.product), ("PH-16", "PILAR"))
        self.assertEqual((pp.piece, pp.product), ("PP-16", "PILAR"))

    def test_truncated_section_last_zero_is_recovered(self) -> None:
        draft = parse_label_text(
            "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: PILAR "
            "Secao: 40X6 Comprimento: 25,140 Peso: 15243,26 "
            "Vol. (m3): 6,005 Peca: PH-12",
            message_date="02/08/2026",
        )
        self.assertEqual(draft.section, "40X60")

    def test_abbreviated_section_label_is_recognized(self) -> None:
        draft = parse_label_text(
            "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: VIGA - VASO "
            "S:54X60 Comprimento: 7,440 Vol. (m3): 1,678 Peca: VPT-2",
            message_date="12/08/2026",
        )
        self.assertEqual(draft.section, "54X60")
        self.assertEqual(draft.warnings, [])

    def test_volume_leading_digit_can_be_recovered_from_weight(self) -> None:
        draft = parse_label_text(
            "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: PILAR "
            "Secao: 40X6 Comprimento: 25,140 Peso (kg): 15243,26 "
            "Vol. (m3): <005 Peca: PH-12",
            message_date="02/08/2026",
        )
        self.assertEqual(draft.section, "40X60")
        self.assertEqual(draft.unit_volume, 6.005)
        self.assertEqual(draft.warnings, [])

    def test_piece_before_label_does_not_infer_product_from_volume_word(self) -> None:
        draft = parse_label_text(
            "MKM GALPAO COMER obra: 0226 - RETANGULAR 30*S0 12,980 "
            "comprimento Peso 4860,10 pp-17 peca vol. 1619585 1,960",
            message_date="11/08/2026",
        )
        self.assertEqual(draft.product, "PILAR")
        self.assertEqual(draft.section, "30X50")
        self.assertEqual(draft.type_name, "PILAR")
        self.assertEqual(draft.warnings, [])

    def test_same_number_is_not_accepted_as_length_and_volume(self) -> None:
        draft = parse_label_text(
            "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: PAINEL "
            "Secao: E=8 Comprimento: 0,899 Vol. (m3): 0,899 Peca: PM-14",
            message_date="11/08/2026",
        )
        self.assertEqual(draft.length, "0,899")
        self.assertIsNone(draft.unit_volume)
        self.assertIn("Confirmar volume unitário", draft.warnings)

    def test_repeated_explicit_length_beats_isolated_shifted_number(self) -> None:
        draft = parse_label_text(
            "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: PAINEL Secao: E=8 "
            "Comprimento (m): 0,899 Peso: 2118,40 Peca: PM-14 "
            "Comprimento (m): 7,490 Comprimento (m): 7,490 "
            "Comprimento (m): 7,490 Vol. (m3): 0,899",
            message_date="11/08/2026",
        )
        self.assertEqual(draft.length, "7,490")
        self.assertEqual(draft.unit_volume, 0.899)
        self.assertEqual(draft.warnings, [])

    def test_extended_piece_section_and_product_catalog_patterns(self) -> None:
        viga = parse_label_text(
            "Obra: AMPL ALUMINIO SJ 40X60 11,025 1,200 VR-3",
            message_date="03/08/2026",
        )
        muro = parse_label_text(
            "Obra: AMPLIACAO ALUMINIO SJ Secao: VARIAVEL Comprimento: 4,500 "
            "Vol. (m3): 0,900 Peca: MA_03-A",
            message_date="03/08/2026",
        )
        laje = parse_label_text(
            "Obra: AMPLIACAO ALUMINIO SJ Produto: LAJE - MACICA Secao: 234X455 "
            "Comprimento: 5,000 Vol. (m3): 2,000 Peca: PL5",
            message_date="03/08/2026",
        )
        self.assertEqual(
            (viga.product, viga.type_name),
            ("VIGA 10,1m ATÉ 15,55m", "VIGA 10,1m ATÉ 15,55m"),
        )
        self.assertEqual((muro.piece, muro.section, muro.product), ("MA_03-A", "VARIAVEL", "MURO"))
        self.assertEqual((laje.product, laje.section), ("LAJE ALVEOLAR", "234X455"))

    def test_overlong_ocr_field_is_not_accepted_as_work(self) -> None:
        draft = parse_label_text(
            "Obra: ROMANEIO DE CARGAS 2301 VENTISOL - GALPAO VERTICAL 282 "
            + "texto ilegivel " * 20,
            message_date="22/07/2026",
        )
        self.assertEqual(draft.work, "VENTISOL - GALPAO VERTICAL 282")

    def test_known_unitermi_work_is_recovered_without_obra_label(self) -> None:
        draft = parse_label_text(
            "UNITERMI - AMPLIACAO FABRIL Produto: PILAR - RETANGULAR "
            "Secao: 50X50 Comprimento: 16,471 Vol. (m3): 4,266 Peca: PP001",
            message_date="12/08/2026",
        )
        self.assertEqual(draft.work, "UNITERMI - AMPLIACAO FABRIL")
        self.assertEqual(draft.warnings, [])

    def test_romaneio_document_creates_one_draft_per_piece(self) -> None:
        drafts = parse_document_text(
            """
            ROMANEIO DE CARGAS | 0104 AMPLIACAO ALUMINIO SJ Obra
            PRELLOG PREFABRICA PRELLOG PREFABRICA PL120 PL124 Total:
            E=10 E=10 11,290 9,520 20,810 m Volume 1,350 1,140 2,490 m3
            Peso 3177,01 2678,80 PAINEL PAINEL MACICO MACICO
            """,
            message_id="romaneio-1",
            message_date="26/07/2026",
        )
        self.assertEqual([item.piece for item in drafts], ["PL120", "PL124"])
        self.assertEqual([item.section for item in drafts], ["E=10", "E=10"])
        self.assertEqual([item.length for item in drafts], ["11,290", "9,520"])
        self.assertEqual([item.unit_volume for item in drafts], [1.35, 1.14])
        self.assertEqual([item.type_name for item in drafts], ["PAINEL", "PAINEL"])

    def test_cejens_porto_imbituba_rule_has_priority(self) -> None:
        self.assertEqual(
            normalize_type("CEJEN - PORTO IMBITUBA", "PAINEL - MACICO", 11.1),
            "METRO CÚBICO",
        )

    def test_viga_terca_uses_only_t_group(self) -> None:
        self.assertEqual(normalize_type("OBRA", "VIGA", 11.5, "TERÇA T"), "VIGA TERÇA")
        self.assertNotEqual(normalize_type("OBRA", "VIGA", 11.5, "TERÇA I"), "VIGA TERÇA")

    def test_suspicious_piece_and_section_are_never_approved_automatically(self) -> None:
        draft = parse_label_text(
            "Obra: AMPLIACAO ALUMINIO SJ Produto: PAINEL - MACICO "
            "Secao: EziO Comprimento (m): 11,110 Peso (kg): 3126,45 "
            "PLI Peca: 1219040 Vol. (m3): 1,330",
            message_date="18/07/2026",
        )
        self.assertEqual(draft.status, "CONFIRMAR")
        self.assertTrue(any("peça" in warning for warning in draft.warnings))
        self.assertTrue(any("seção" in warning for warning in draft.warnings))


class GroupingTests(unittest.TestCase):
    def _draft(self, date_value: str, piece: str = "PL1", dimensions: str = "E=10 11,110"):
        draft = parse_label_text(
            "Obra: AMPLIACAO ALUMINIO SJ Produto: PAINEL - MACICO "
            "Secao: E=10 Comprimento (m): 11,110 Peso: 1 Peca: PL1 Vol. (m3): 1,330",
            message_date=date_value,
        )
        draft.piece = piece
        draft.dimensions = dimensions
        draft.status = "APROVADO"
        draft.warnings = []
        return draft

    def test_identical_approved_pieces_from_same_day_are_grouped(self) -> None:
        rows = group_approved_drafts([self._draft("18/07/2026"), self._draft("18/07/2026")])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].quantity, 2)
        self.assertEqual(rows[0].source_count, 2)

    def test_different_days_are_never_grouped(self) -> None:
        rows = group_approved_drafts([self._draft("18/07/2026"), self._draft("19/07/2026")])
        self.assertEqual(len(rows), 2)

    def test_any_different_field_creates_another_row(self) -> None:
        rows = group_approved_drafts([
            self._draft("18/07/2026"),
            self._draft("18/07/2026", dimensions="E=10 11,210"),
        ])
        self.assertEqual(len(rows), 2)

    def test_unapproved_drafts_are_excluded(self) -> None:
        draft = self._draft("18/07/2026")
        draft.status = "CONFIRMAR"
        self.assertEqual(group_approved_drafts([draft]), [])


class ReviewConsensusTests(unittest.TestCase):
    def test_duplicate_token_after_volume_keeps_volume_only(self) -> None:
        draft = parse_label_text(
            "Obra: AMPLIACAO ALUMINIO SJ Produto: PAINEL Secao: E=10 "
            "Peca: PL109 Vol. (m3): 1,370",
            message_date="06/08/2026",
        )
        draft.length = "1,370"
        draft.unit_volume = 1.37
        draft.ocr_text = "comprimento Vol. (m3): 1,370"

        _sanitize_duplicate_measurements(draft)

        self.assertEqual(draft.length, "")
        self.assertEqual(draft.unit_volume, 1.37)
        self.assertIn("Confirmar comprimento", draft.warnings)

    def test_duplicate_token_after_length_keeps_length_only(self) -> None:
        draft = parse_label_text(
            "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: VIGA Secao: 25X80 "
            "Comprimento: 7,180 Peca: VL-3",
            message_date="11/08/2026",
        )
        draft.length = "7,180"
        draft.unit_volume = 7.18
        draft.ocr_text = "Comengnento (m): 7,180 Peso 3040,70 Vol."

        _sanitize_duplicate_measurements(draft)

        self.assertEqual(draft.length, "7,180")
        self.assertIsNone(draft.unit_volume)
        self.assertIn("Confirmar volume unitário", draft.warnings)

    def test_strong_majority_only_fills_missing_length(self) -> None:
        drafts = []
        for length, volume in (("11,534", 1.37), ("11,534", 1.37), ("11,534", 1.37), ("1,370", 1.37), ("", 1.37)):
            draft = parse_label_text(
                "Obra: AMPLIACAO ALUMINIO SJ Produto: PAINEL Secao: E=10 "
                "Peca: PL109 Vol. (m3): 1,370",
                message_date="06/08/2026",
            )
            draft.length = length
            draft.unit_volume = volume
            draft.warnings = ["Confirmar comprimento"] if not length else []
            drafts.append(draft)

        _apply_message_consensus(drafts)

        self.assertEqual(drafts[-1].length, "11,534")
        self.assertEqual(drafts[3].length, "11,534")

    def test_weak_majority_does_not_fill_missing_value(self) -> None:
        drafts = []
        for length in ("11,534", "11,534", "9,520", ""):
            draft = parse_label_text(
                "Obra: AMPLIACAO ALUMINIO SJ Produto: PAINEL Secao: E=10 "
                "Peca: PL109 Vol. (m3): 1,370",
                message_date="06/08/2026",
            )
            draft.length = length
            draft.warnings = ["Confirmar comprimento"] if not length else []
            drafts.append(draft)

        _apply_message_consensus(drafts)

        self.assertEqual(drafts[-1].length, "")

    def test_reverse_fingerprint_tolerates_one_confused_piece_code(self) -> None:
        drafts = []
        for piece in ("PM-4", "PM-4", "PM-4", "PM-4", "PM-14", ""):
            draft = parse_label_text(
                "Obra: VENTISOL - GALPAO VERTICAL 282 Produto: PAINEL "
                "Secao: E=8 Comprimento: 7,490 Vol. (m3): 0,899",
                message_date="11/08/2026",
            )
            draft.piece = piece
            draft.warnings = ["Confirmar peça"] if not piece else []
            drafts.append(draft)

        _apply_message_consensus(drafts)

        self.assertEqual(drafts[-1].piece, "PM-4")
        self.assertEqual(drafts[4].piece, "PM-14")

    def test_historical_piece_catalog_fills_missing_field_with_strong_majority(self) -> None:
        drafts = []
        for index, length in enumerate(("11,534",) * 6 + ("1,534",) + ("",)):
            draft = parse_label_text(
                "Obra: AMPLIACAO ALUMINIO SJ Produto: PAINEL Secao: E=10 "
                "Peca: PL109 Vol. (m3): 1,370",
                message_date=f"{index + 1:02d}/08/2026",
            )
            draft.length = length
            draft.warnings = ["Confirmar comprimento"] if not length else []
            drafts.append(draft)

        _apply_message_consensus(drafts)

        self.assertEqual(drafts[-1].length, "11,534")
        self.assertEqual(drafts[6].length, "1,534")


class WorkbookTests(unittest.TestCase):
    def test_approved_row_payload_keeps_only_source_fields(self) -> None:
        payload = rows_to_payload([ConsolidatedRow(
            type_name="PAINEL",
            message_date="24/07/2026",
            work="VENTISOL",
            quantity=2,
            piece="PL-14-B",
            dimensions="E=8 7,350",
            unit_volume=0.419,
            cargo_type="PEÇAS ESTOQUE",
            source_count=2,
        )])[0]
        self.assertEqual(payload["message_date"], "2026-07-24")
        self.assertEqual(payload["quantity"], 2)
        self.assertNotIn("volume_total", payload)
        self.assertNotIn("unit_measure", payload)
        self.assertNotIn("unit_price", payload)
        self.assertNotIn("total_price", payload)

    def _create_valid_workbook(self, path: Path) -> None:
        workbook = Workbook()
        target = workbook.active
        target.title = "2ª quinz.julho"
        target.append(["Medição AWL"])
        for _ in range(4):
            target.append([])
        target.append([
            "TIPO", "DATA", "OBRA", "QUANT. (PÇ ou m)", "PEÇA", None,
            "DIMENSÕES", "VOL UNIT. (m³)", "VOL TOTAL (m³)",
            "TIPO DE CARGA", "UNID. DE MEDIDA", "R$ UNIT.", "R$ TOTAL",
        ])
        target["D20"] = "=SUM(D7:D19)"
        table = workbook.create_sheet("TABELA")
        table.append(["TIPO", "VALOR", "UNIDADE"])
        table.append(["ESTACA", 0.75, "m"])
        workbook.save(path)

    def test_valid_workbook_is_accepted(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "medicoes.xlsx"
            self._create_valid_workbook(path)
            result = validate_workbook(path, PeriodSelection(2026, 7, 2))
            self.assertTrue(result.valid, result.messages)
            self.assertEqual(result.details["header_row"], 6)

    def test_missing_target_sheet_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "medicoes.xlsx"
            self._create_valid_workbook(path)
            result = validate_workbook(path, PeriodSelection(2026, 6, 2))
            self.assertFalse(result.valid)

    def test_backup_is_created_in_expected_folder(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "medicoes.xlsx"
            source.write_bytes(b"example")
            backup = create_backup(source, datetime(2026, 7, 26, 10, 30, 0))
            self.assertTrue(backup.exists())
            self.assertEqual(backup.parent.name, "Backups Fechamento")
            self.assertEqual(backup.read_bytes(), b"example")

    def test_writer_imports_into_copy_with_backup_and_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "medicoes.xlsx"
            self._create_valid_workbook(path)
            row = ConsolidatedRow(
                type_name="PILAR",
                message_date="31/07/2026",
                work="OBRA TESTE",
                quantity=2,
                piece="PP-10",
                dimensions="40X40 10,000",
                unit_volume=1.25,
                cargo_type="PEÇAS ESTOQUE",
                source_count=2,
            )
            result = write_approved_rows(
                path, PeriodSelection(2026, 7, 2), [row]
            )
            self.assertEqual(result.imported_rows, [7])
            self.assertTrue(result.backup_path.exists())
            workbook = load_workbook(path, data_only=False)
            sheet = workbook["2ª quinz.julho"]
            self.assertEqual(sheet["A7"].value, "PILAR")
            self.assertEqual(sheet["E7"].value, "PP")
            self.assertEqual(sheet["F7"].value, "10")
            self.assertEqual(sheet["I7"].value, "=H7*D7")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
