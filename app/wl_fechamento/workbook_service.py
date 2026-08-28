from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from .models import PeriodSelection, WorkbookValidation


REQUIRED_HEADERS = {"TIPO", "DATA", "OBRA"}
QUANTITY_HEADER_PREFIX = "QUANT"
REQUIRED_TABLE_SHEET = "TABELA"
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    return re.sub(r"\s+", " ", text)


def _find_header_row(worksheet, max_rows: int = 15) -> tuple[int | None, list[str]]:
    worksheet_max_row = worksheet.max_row or 0
    for row_number in range(1, min(worksheet_max_row, max_rows) + 1):
        values = [_normalize(cell.value) for cell in worksheet[row_number]]
        present = set(values)
        has_quantity = any(value.startswith(QUANTITY_HEADER_PREFIX) for value in values)
        if REQUIRED_HEADERS.issubset(present) and has_quantity:
            return row_number, values
    return None, []


def validate_workbook(path: str | Path, period: PeriodSelection) -> WorkbookValidation:
    # Importação tardia: mantém a abertura da janela rápida e só carrega a
    # biblioteca do Excel quando Lilian solicita a validação.
    from openpyxl import load_workbook

    workbook_path = Path(path).expanduser().resolve()
    result = WorkbookValidation(
        path=workbook_path,
        valid=False,
        target_sheet=period.sheet_name,
    )

    if not workbook_path.exists() or not workbook_path.is_file():
        result.messages.append("O arquivo selecionado não existe.")
        return result
    if workbook_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        result.messages.append("Selecione uma planilha .xlsx ou .xlsm.")
        return result

    try:
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=False,
            keep_vba=workbook_path.suffix.lower() == ".xlsm",
        )
    except Exception as exc:  # openpyxl fornece diferentes erros por arquivo
        result.messages.append(f"Não foi possível abrir a planilha: {exc}")
        return result

    try:
        result.details["sheet_names"] = list(workbook.sheetnames)
        if REQUIRED_TABLE_SHEET not in workbook.sheetnames:
            result.messages.append("A aba TABELA não foi encontrada.")
        if period.sheet_name not in workbook.sheetnames:
            result.messages.append(
                f"A aba de destino '{period.sheet_name}' não foi encontrada."
            )
        else:
            target = workbook[period.sheet_name]
            header_row, headers = _find_header_row(target)
            if header_row is None:
                result.messages.append(
                    "A aba de destino não contém os cabeçalhos obrigatórios "
                    "TIPO, DATA, OBRA e QUANTIDADE."
                )
            else:
                result.details["header_row"] = header_row
                result.details["headers"] = headers
                result.messages.append(
                    f"Aba '{period.sheet_name}' validada; cabeçalhos na linha {header_row}."
                )

        if REQUIRED_TABLE_SHEET in workbook.sheetnames:
            table_sheet = workbook[REQUIRED_TABLE_SHEET]
            table_max_row = table_sheet.max_row or 0
            if table_max_row < 2:
                result.messages.append("A aba TABELA está vazia.")
            else:
                result.details["table_rows"] = table_max_row

        result.valid = not any(
            message.startswith("A aba TABELA não")
            or message.startswith("A aba de destino '")
            or message.startswith("A aba de destino não")
            or message.startswith("A aba TABELA está")
            for message in result.messages
        )
        if result.valid:
            result.messages.insert(0, "Planilha válida e pronta para configuração.")
    finally:
        workbook.close()

    return result


def create_backup(path: str | Path, now: datetime | None = None) -> Path:
    """Cria uma cópia completa antes da futura importação."""
    source = Path(path).expanduser().resolve()
    if not source.exists() or not source.is_file():
        raise FileNotFoundError("A planilha oficial não foi encontrada.")
    timestamp = (now or datetime.now()).strftime("%Y-%m-%d_%H-%M-%S")
    backup_directory = source.parent / "Backups Fechamento"
    backup_directory.mkdir(parents=True, exist_ok=True)
    destination = backup_directory / f"{source.stem}_backup_{timestamp}{source.suffix}"
    shutil.copy2(source, destination)
    return destination
