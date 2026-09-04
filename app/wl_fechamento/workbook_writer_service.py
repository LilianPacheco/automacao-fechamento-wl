from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from .grouping_service import ConsolidatedRow
from .models import PeriodSelection
from .workbook_service import create_backup, validate_workbook


@dataclass(frozen=True)
class WorkbookWriteResult:
    workbook_path: Path
    backup_path: Path
    imported_rows: list[int]


def _iso_date(value: str) -> str:
    return datetime.strptime(value, "%d/%m/%Y").date().isoformat()


def rows_to_payload(rows: list[ConsolidatedRow]) -> list[dict[str, object]]:
    return [
        {
            "status": "APROVADO",
            "message_date": _iso_date(row.message_date),
            "type_name": row.type_name,
            "work": row.work,
            "quantity": row.quantity,
            "piece": row.piece,
            "dimensions": row.dimensions,
            "unit_volume": row.unit_volume,
            "cargo_type": row.cargo_type,
        }
        for row in rows
    ]


def _split_piece(value: str) -> tuple[str, str]:
    compact = re.sub(r"\s+", "", str(value or ""))
    match = re.match(r"^([A-Za-zÀ-ÿ]+)-?(.*)$", compact)
    if not match:
        return compact, ""
    return match.group(1).rstrip("-"), match.group(2).lstrip("-")


def _fallback_openpyxl_write(
    source: Path,
    temporary_output: Path,
    period: PeriodSelection,
    rows: list[ConsolidatedRow],
    header_row: int,
) -> list[int]:
    """Fallback for workbooks whose formulas/styles are not round-tripped by the JS exporter.

    The original file is never opened for writing: a temporary copy is edited,
    validated, and only then promoted by the caller.
    """
    from openpyxl import load_workbook

    import shutil
    shutil.copy2(source, temporary_output)
    workbook = load_workbook(temporary_output, data_only=False, keep_vba=temporary_output.suffix.lower() == ".xlsm")
    sheet = workbook[period.sheet_name]
    total_row = None
    for row_number in range(header_row + 1, (sheet.max_row or 0) + 1):
        formulas = [sheet.cell(row_number, col).value for col in range(1, 14)]
        if any(isinstance(value, str) and "SUM(" in value.upper() for value in formulas):
            total_row = row_number
            break
    if total_row is None:
        raise RuntimeError("A linha de total não foi localizada no arquivo oficial.")

    free_rows = []
    for row_number in range(header_row + 1, total_row):
        source_values = [sheet.cell(row_number, col).value for col in [1, 2, 3, 4, 5, 6, 7, 10]]
        if all(value in (None, "") for value in source_values):
            free_rows.append(row_number)
        if len(free_rows) >= len(rows):
            break
    if len(free_rows) < len(rows):
        raise RuntimeError("A aba oficial não possui linhas livres suficientes.")

    imported_rows = []
    for item, target_row in zip(rows, free_rows):
        piece_prefix, piece_number = _split_piece(item.piece)
        values = {
            1: item.type_name,
            2: datetime.strptime(item.message_date, "%d/%m/%Y"),
            3: item.work,
            4: item.quantity,
            5: piece_prefix,
            6: piece_number,
            7: item.dimensions,
            8: item.unit_volume,
            10: item.cargo_type,
        }
        for col, value in values.items():
            sheet.cell(target_row, col).value = value
        sheet.cell(target_row, 9).value = f"=H{target_row}*D{target_row}"
        sheet.cell(target_row, 11).value = f"=VLOOKUP(A{target_row},'TABELA'!$A$2:$C$200,3,0)"
        sheet.cell(target_row, 12).value = f"=VLOOKUP(A{target_row},'TABELA'!$A$2:$C$200,2,0)"
        sheet.cell(target_row, 13).value = (
            f'=IF($K{target_row}="PÇ",$L{target_row}*$D{target_row},'
            f'IF($K{target_row}="m³",$I{target_row}*$L{target_row},$D{target_row}*$L{target_row}))'
        )
        sheet.cell(target_row, 2).number_format = "dd/mm/yyyy"
        sheet.cell(target_row, 8).number_format = "0.000"
        sheet.cell(target_row, 9).number_format = "0.000"
        imported_rows.append(target_row)
    workbook.save(temporary_output)
    workbook.close()
    return imported_rows


def write_approved_rows(
    workbook_path: str | Path,
    period: PeriodSelection,
    rows: list[ConsolidatedRow],
) -> WorkbookWriteResult:
    if not rows:
        raise ValueError("Não há linhas aprovadas para importar.")
    source = Path(workbook_path).expanduser().resolve()
    validation = validate_workbook(source, period)
    if not validation.valid:
        raise RuntimeError("A planilha deixou de ser válida; nenhuma escrita foi feita.")
    backup = create_backup(source)
    temporary_output = source.with_name(f".{source.stem}.wl-{uuid.uuid4().hex}.xlsx")
    try:
        imported_rows = _fallback_openpyxl_write(
            source,
            temporary_output,
            period,
            rows,
            int(validation.details["header_row"]),
        )
        prepared_validation = validate_workbook(temporary_output, period)
    except Exception:
        temporary_output.unlink(missing_ok=True)
        raise
    if not prepared_validation.valid:
        temporary_output.unlink(missing_ok=True)
        raise RuntimeError(
            "A cópia preparada não passou na validação. O arquivo oficial não foi alterado."
        )

    try:
        temporary_output.replace(source)
    except OSError as exc:
        raise RuntimeError(
            "A cópia foi preparada, mas o arquivo está aberto no Excel. "
            "Feche a planilha e tente novamente; o backup foi preservado."
        ) from exc
    return WorkbookWriteResult(source, backup, imported_rows)
